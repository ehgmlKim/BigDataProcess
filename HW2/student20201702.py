#!/usr/bin/python3
import openpyxl
wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']



rank_list=[] # 등수 넣은 배열
rank_rowid = {} # row_id : rank
rank_dict={} #등수에 해당하는 학생 수 rank : num

row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1

row_id = 1
for row in ws:
	rank=1
	row_id2=1
	if row_id != 1:
		for row in ws:
			if row_id2 != 1:
				a=ws.cell(row = row_id, column=7).value
				b=ws.cell(row = row_id2, column=7).value
				if a < b:
					rank += 1
				ws.cell(row=row_id, column=8).value = rank
				rank_rowid[row_id] = rank
			row_id2 += 1
	row_id +=1

total = row_id -2 #전체 학생 수
for i in range(1, row_id):
	if i != 1:
		rank_list.append(ws.cell(row=i, column=8).value)


for e in range(len(rank_list)):
	if rank_list[e] not in rank_dict:
		rank_dict[rank_list[e]] = 1
	else :
		rank_dict[rank_list[e]] += 1
	row_id += 1




Rid = list(rank_rowid.items())
Rid.sort(key = lambda x:x[1]) # row_id : rank
#print("Rid :", Rid)
Rnum= list(rank_dict.items()) #만든 딕셔너리를 리스트로 변환 등수 : 등수에 해당하는 학생
Rnum.sort(key=lambda x:x[0])
#print("Rnum : ", Rnum)
getA=[]
getAp=[]
getB=[]
getBp=[]
getC=[]
getCp=[]


#A,B,C 등급 먼저 알아내기
for k in range(len(Rnum)):
	count= Rnum[k][1]
	rank = Rnum[k][0]
	if rank <= total*0.3 and len(getA) + count <= total * 0.3:
		for i in range(len(Rid)):
			if rank == Rid[i][1] :
				getA.append(Rid[i][0])
	elif rank <= total * 0.7 and len(getB) + len(getA) + count  <= total * 0.7:
		for i in range(len(Rid)):
			if rank == Rid[i][1] :
				getB.append(Rid[i][0])
	elif rank > total*0.7:
		for i in range(len(Rid)):
			if rank == Rid[i][1] :
				getC.append(Rid[i][0])
	else:
		pass


# + 붙일 애들 알아내기
for k in range(len(Rnum)):
	count= Rnum[k][1]
	rank = Rnum[k][0]
	rowid=[]
	for f in range(len(Rid)):
		if rank == Rid[f][1]:
			rowid.append(Rid[f][0])
	if len(getAp) + count <= len(getA) * 0.5:
		for r in rowid:
			if r in getA:
				getAp.append(r)
	if len(getBp) + count <= len(getB) * 0.5:
		for r in rowid:
			if r in getB:
				getBp.append(r)
	if len(getCp) + count <= len(getC) * 0.5:
		for r in rowid:
			if r in getC:
				getCp.append(r)
	else:
		pass	
#점수 부여하기	
for k in getA:	
	ws.cell(row=k, column=8).value = 'A0'
for k in getB:
	ws.cell(row=k, column=8).value = 'B0'
for k in getC:
	ws.cell(row=k, column=8).value = 'C0'
for k in getAp:	
	ws.cell(row=k, column=8).value = 'A+'
for k in getBp:
	ws.cell(row=k, column=8).value = 'B+'
for k in getCp:
	ws.cell(row=k, column=8).value = 'C+'
	
	




wb.save("student.xlsx")















